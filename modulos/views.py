from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User, Group
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError, transaction
from django.contrib.auth.decorators import login_required
from .models import Producto, Insumo, Proveedor, Cliente, Ordendeproduccion, Venta, Metododepago, Usuario, Detalleventa, Compraproducto, Comprainsumo, Controlgastoinsumo
from .forms import ProveedorForm, ClienteForm, InsumoForm, ProductoForm, OrdendeproduccionForm, ClienteSignUpForm, VentaForm, DetalleVentaForm, CompraProductoForm, CompraInsumoForm, VentaClienteForm, GastoInsumoForm
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.core.mail import send_mail, BadHeaderError
from django.contrib import messages
from django.db.models import F, Sum
import csv 
import matplotlib.pyplot as plt
import io
import base64






from django.contrib.auth.decorators import user_passes_test

# Función que verifica si el usuario pertenece al grupo 'Usuarios'

def es_usuario(user):
    return user.groups.filter(name__in=['Usuarios', 'Administrador', 'Jefe de Producción']).exists()




def productos(request):
    # Obtener todos los productos desde la base de datos
    productos = Producto.objects.all()
    return render(request, 'productos.html', {'productos': productos})


# views.py

from django.shortcuts import render
from .models import Producto

def catalogos(request):
    productos = Producto.objects.all()

    # Filtros obtenidos de la URL
    categoria = request.GET.get('categoria')
    genero = request.GET.get('genero')  # Aquí recibimos 'masculino' o 'femenino'
    color = request.GET.get('color')
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    nombre = request.GET.get('nombre')

    # Aplicar los filtros si existen
    if categoria:
        productos = productos.filter(nombrecategoria=categoria)
    
    if genero:
        productos = productos.filter(genero=genero)  # El filtro ahora aplicará 'masculino' o 'femenino'

    if color:
        productos = productos.filter(color=color)

    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    
    if nombre:
        productos = productos.filter(nombreproducto__icontains=nombre)

    # Carrito de compras
    carrito = request.session.get('carrito', {})
    total = calcular_total_carrito(carrito)

    return render(request, 'catalogo.html', {'productos': productos, 'carrito': carrito, 'total': total})



from django.http import JsonResponse


def agregar_al_carrito(request, producto_id):
    producto = get_object_or_404(Producto, idproducto=producto_id)
    
    # Obtener el carrito de la sesión, o crear un carrito vacío si no existe
    carrito = request.session.get('carrito', {})
    
    # Convertir producto_id a string para evitar posibles problemas de tipo
    producto_id_str = str(producto_id)
    
    # Si el producto no está en el carrito, lo añadimos
    if producto_id_str not in carrito:
        carrito[producto_id_str] = {
            'nombre': producto.nombreproducto,
            'precio': str(producto.precio),  # Convertir a string para la sesión
            'cantidad': 1,
            'imagen': producto.imagenproducto.url if producto.imagenproducto else None  # URL de la imagen
        }
    else:
        # Si el producto ya está en el carrito, incrementar la cantidad
        carrito[producto_id_str]['cantidad'] += 1
    
    # Guardar el carrito actualizado en la sesión
    request.session['carrito'] = carrito
    
    # Redireccionar a la página de productos o donde quieras
    return redirect('catalogos')  # Cambia 'lista_productos' por la vista adecuada







from django.shortcuts import render

def mostrar_carrito(request):
    carrito = request.session.get('carrito', {})
    total = calcular_total_carrito(carrito)  # Calcula el total del carrito
    return render(request, 'carrito.html', {'carrito': carrito, 'total': total})


def calcular_total_carrito(carrito):
    total = 0
    for key, item in carrito.items():
        # Convertimos el precio a float antes de multiplicar
        precio = float(item['precio'])
        total += item['cantidad'] * precio
    return total



from django.shortcuts import redirect

# Ejemplo de la función actualizar_cantidad_carrito
# views.py
def actualizar_cantidad_carrito(request, producto_id, accion):
    carrito = request.session.get('carrito', {})

    # Verifica si el producto está en el carrito
    if str(producto_id) in carrito:
        if accion == 'incrementar':
            carrito[str(producto_id)]['cantidad'] += 1
        elif accion == 'disminuir' and carrito[str(producto_id)]['cantidad'] > 1:
            carrito[str(producto_id)]['cantidad'] -= 1

    # Actualiza el carrito en la sesión
    request.session['carrito'] = carrito

    # Redirigir a la página anterior
    previous_page = request.META.get('HTTP_REFERER', 'catalogos')
    return redirect(previous_page)







# Ejemplo de la función eliminar_del_carrito
# views.py
def eliminar_del_carrito(request, producto_id):
    carrito = request.session.get('carrito', {})

    # Verifica que el producto exista en el carrito y elimínalo
    if str(producto_id) in carrito:
        del carrito[str(producto_id)]

    # Actualiza el carrito en la sesión
    request.session['carrito'] = carrito

    # Redirigir a la página anterior (carrito o catálogo)
    previous_page = request.META.get('HTTP_REFERER', 'mostrar_carrito')
    return redirect(previous_page)



from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.contrib import messages


@login_required
def finalizar_compra(request):
    carrito = request.session.get('carrito', {})

    if not carrito:
        return redirect('mostrar_carrito')

    # Buscar el cliente asociado al usuario
    cliente = get_object_or_404(Cliente, user=request.user)

    # Inicializar variables para `productos_carrito` y `total_venta`
    productos_carrito = []
    total_venta = 0

    if request.method == 'POST':
        print("POST recibido")
        venta_form = VentaClienteForm(request.POST)

        if venta_form.is_valid():
            print("Formulario de venta válido")
            venta = venta_form.save(commit=False)
            venta.cliente_idclientes = cliente
            venta.valortotal = 0  # Se calcula más adelante
            venta.modalidadventa = 'virtual'  # Siempre será virtual
            venta.save()

            total = 0
            detalles_venta = []  # Lista para almacenar los detalles de venta

            # Procesar los productos en el carrito y guardarlos en DetalleVenta
            for producto_id, item in carrito.items():
                producto = Producto.objects.get(idproducto=producto_id)
                cantidad = item['cantidad']
                precio = float(item['precio'])
                subtotal = cantidad * precio

                # Crear y guardar DetalleVenta
                detalle_venta = Detalleventa(
                    producto_idproducto=producto,
                    venta_idventas=venta,
                    cantidad=cantidad,
                    subtotal=subtotal
                )
                detalle_venta.save()

                # Agregar el detalle a la lista de detalles de venta
                detalles_venta.append(detalle_venta)

                # Restar la cantidad comprada del inventario del producto
                if producto.cantidad >= cantidad:
                    producto.cantidad -= cantidad
                    producto.save()
                else:
                    messages.error(request, f'No hay suficiente stock para {producto.nombreproducto}.')
                    return redirect('mostrar_carrito')

                total += subtotal

            # Guardar el valor total de la venta
            venta.valortotal = total
            venta.save()

            # Enviar el correo de confirmación de la compra
            try:
                asunto = 'Confirmación de compra'
                # Renderizar el contenido del correo usando una plantilla
                mensaje_html = render_to_string('clientes_pagina/ventas_correo.html', {
                    'cliente': cliente,
                    'venta': venta,
                    'detalles_venta': detalles_venta,  # Aquí pasamos la lista de detalles de venta
                    'total': total
                })
                mensaje_texto = strip_tags(mensaje_html)
                destinatario = cliente.user.email

                send_mail(
                    asunto,
                    mensaje_texto,  # El mensaje en texto plano
                    settings.DEFAULT_FROM_EMAIL,  # El remitente
                    [destinatario],  # Lista de destinatarios
                    html_message=mensaje_html  # El mensaje en HTML
                )

                messages.success(request, '¡Compra finalizada con éxito! Se ha enviado un correo de confirmación.')

            except Exception as e:
                messages.error(request, f'Ocurrió un error al enviar el correo: {str(e)}')

            # Limpiar el carrito después de la compra
            request.session['carrito'] = {}

            return redirect('confirmacion_compra')
        else:
            print("Formulario de venta no válido:", venta_form.errors)

            # En caso de formulario inválido, reconstruimos `productos_carrito` y `total_venta`
            for producto_id, item in carrito.items():
                producto = Producto.objects.get(idproducto=producto_id)
                cantidad = item['cantidad']
                precio = float(item['precio'])
                subtotal = cantidad * precio

                productos_carrito.append({
                    'nombre': item['nombre'],
                    'cantidad': cantidad,
                    'precio': precio,
                    'subtotal': subtotal,
                })

                total_venta += subtotal

    else:
        # Inicializamos el formulario de venta para solicitudes GET
        venta_form = VentaClienteForm(cliente=cliente)

        # En caso de solicitud GET, preparamos `productos_carrito`
        for producto_id, item in carrito.items():
            producto = Producto.objects.get(idproducto=producto_id)
            cantidad = item['cantidad']
            precio = float(item['precio'])
            subtotal = cantidad * precio

            productos_carrito.append({
                'nombre': item['nombre'],
                'cantidad': cantidad,
                'precio': precio,
                'subtotal': subtotal,
            })

            total_venta += subtotal

    return render(request, 'ventas/finalizar_compra.html', {
        'venta_form': venta_form,
        'productos_carrito': productos_carrito,
        'total_venta': total_venta,
        'venta_form_errors': venta_form.errors if not venta_form.is_valid() else None,
    })



from django.shortcuts import render

def confirmacion_compra(request):
    return render(request, 'ventas/confirmacion_compra.html')





# Venta Nathalia 



from django.forms import modelformset_factory
from .models import Detalleventa
from .forms import VentaForm, DetalleVentaFormSet
from django.contrib.auth.decorators import permission_required
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings

@permission_required('modulos.add_venta', login_url='inicio')
@login_required
def crear_venta(request, cliente_id):
    cliente = get_object_or_404(Cliente, idclientes=cliente_id)

    if request.method == 'POST':
        # Formularios de venta y formset de detalle de venta
        venta_form = VentaForm(request.POST, cliente=cliente)
        formset = DetalleVentaFormSet(request.POST, queryset=Detalleventa.objects.none())

        if venta_form.is_valid() and formset.is_valid():
            productos_seleccionados = False  # Para verificar si se seleccionó algún producto

            # Validar si hay algún producto seleccionado
            for form in formset:
                if form.cleaned_data and form.cleaned_data.get('producto_idproducto'):
                    productos_seleccionados = True  # Al menos un producto está seleccionado
                    break  # Ya encontramos un producto seleccionado, no es necesario seguir

            if not productos_seleccionados:
                messages.error(request, 'Debe seleccionar al menos un producto.')
                return render(request, 'ventas/formulario_ventas.html', {
                    'venta_form': venta_form,
                    'formset': formset,
                    'cliente': cliente
                })

            # Guardamos la venta sin valor total inicialmente
            venta = venta_form.save(commit=False)
            venta.cliente_idclientes = cliente
            venta.valortotal = 0  # Inicializamos el valor total temporalmente
            total = 0  # Inicializamos el total general de la venta

            # Validar el inventario primero, antes de guardar cualquier cosa
            for form in formset:
                if form.cleaned_data:
                    producto = form.cleaned_data['producto_idproducto']
                    cantidad = form.cleaned_data['cantidad']
                    if producto.cantidad < cantidad:
                        messages.error(request, f'No hay suficiente stock para {producto.nombreproducto}.')
                        return render(request, 'ventas/formulario_ventas.html', {
                            'venta_form': venta_form,
                            'formset': formset,
                            'cliente': cliente
                        })

            # Si todo está bien, guardamos la venta y los detalles
            venta.save()  # Guardamos la venta para obtener su ID

            for form in formset:
                if form.cleaned_data:
                    detalle_venta = form.save(commit=False)
                    producto = form.cleaned_data['producto_idproducto']
                    cantidad = form.cleaned_data['cantidad']

                    detalle_venta.venta_idventas = venta
                    detalle_venta.subtotal = cantidad * producto.precio

                    # Actualizamos el inventario
                    producto.cantidad -= cantidad
                    producto.save()

                    detalle_venta.save()  # Guardamos el detalle de la venta
                    total += detalle_venta.subtotal

            # Ahora actualizamos el total de la venta
            venta.valortotal = total
            venta.save()  # Guardamos el valor total en la venta

            # **Enviar el correo**
            detalles_venta = Detalleventa.objects.filter(venta_idventas=venta)
            subject = 'Gracias por tu compra!'
            message = render_to_string('clientes_pagina/ventas_correo.html', {
                'venta': venta,
                'detalles_venta': detalles_venta
            })
            recipient_email = cliente.correo1

            try:
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [recipient_email],
                    fail_silently=False,
                    html_message=message
                )
                print("Correo enviado con éxito.")
            except Exception as e:
                print(f"Error enviando el correo: {str(e)}")

            return redirect('resumen_ventas')

    else:
        venta_form = VentaForm(cliente=cliente)
        formset = DetalleVentaFormSet(queryset=Detalleventa.objects.none())

    return render(request, 'ventas/formulario_ventas.html', {
        'venta_form': venta_form,
        'formset': formset,
        'cliente': cliente
    })





from django.http import JsonResponse
from .models import Producto

def obtener_precio_producto(request, producto_id):
    try:
        producto = Producto.objects.get(idproducto=producto_id)
        return JsonResponse({'precio': producto.precio})
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)





# procedimiento almacenado de valor unitario * cantidad, suma de subtotales y actualizacion prodcutos por compra





# Vista para el registro de nuevos usuarios con correo electrónico
def signup(request):
    if request.method == 'GET':
        form = UserCreationForm()
        return render(request, 'signup.html', {'form': form})

    else:
        # Verificar que las contraseñas coincidan
        if request.POST['password1'] == request.POST['password2']:
            try:
                # Crear el usuario usando el correo como 'username' y 'email'
                user = User.objects.create_user(
                    username=request.POST['email'],  # Usamos el email como nombre de usuario
                    email=request.POST['email'],     # También lo guardamos en el campo de correo
                    password=request.POST['password1']
                )
                user.save()

                # Asignar al grupo "Clientes"
                grupo_clientes, created = Group.objects.get_or_create(name='Clientes')
                user.groups.add(grupo_clientes)

                # Autenticar e iniciar sesión automáticamente
                login(request, user)

                # Redirigir según el grupo
                return redirect('cliente_registrado')

            except IntegrityError:
                return render(request, 'signup.html', {
                    'form': UserCreationForm(),
                    'error': 'Ya existe un usuario registrado con este correo.'
                })
        else:
            return render(request, 'signup.html', {
                'form': UserCreationForm(),
                'error': 'Las contraseñas no coinciden.'
            })




from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User, Group
from django.shortcuts import render, redirect

def signin(request):
    if request.method == 'GET':
        form = AuthenticationForm()
        return render(request, 'signin.html', {'form': form})

    else:
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username_or_email = form.cleaned_data.get('username')  # Usamos este campo para email o username
            password = form.cleaned_data.get('password')

            user = None
            try:
                if '@' in username_or_email:
                    # Si es un correo electrónico, usamos filter para obtener todos los usuarios con ese correo
                    users = User.objects.filter(email=username_or_email)

                    if users.exists():
                        # Intentamos autenticar con cada usuario hasta encontrar el correcto
                        for user_by_email in users:
                            user = authenticate(request, username=user_by_email.username, password=password)
                            if user is not None:
                                break  # Usuario autenticado, salimos del ciclo
                else:
                    # Si no es un correo, intentamos autenticar con el nombre de usuario directamente
                    user = authenticate(request, username=username_or_email, password=password)

                if user is not None:
                    login(request, user)

                    # Identificar el grupo al que pertenece el usuario y redirigir
                    if user.groups.filter(name='Administrador').exists():
                        return redirect('inicio')  # Redirigir a la vista de admin
                    elif user.groups.filter(name='Jefe de Producción').exists():
                        return redirect('inicio')  # Redirigir a la vista de jefe de producción
                    elif user.groups.filter(name='Clientes').exists():
                        return redirect('cliente_registrado')  # Redirigir a la vista de clientes
                    else:
                        return redirect('inicio')  # Redirigir a una vista general si no pertenece a ningún grupo

                else:
                    return render(request, 'signin.html', {
                        'form': form,
                        'error': 'Correo, usuario o contraseña incorrectos.'
                    })

            except User.DoesNotExist:
                return render(request, 'signin.html', {
                    'form': form,
                    'error': 'Correo, usuario o contraseña incorrectos.'
                })

        return render(request, 'signin.html', {
            'form': form,
            'error': 'Correo, usuario o contraseña incorrectos.'
        })


def signout(request):
    logout(request)
    return redirect('home')


"""
# Señales para crear y guardar automáticamente el modelo Usuario cuando se crea un User
@receiver(post_save, sender=User)
def create_usuario(sender, instance, created, **kwargs):
    if created:
        Usuario.objects.create(user=instance)

# Asegúrate de que no necesitas actualizar el objeto Usuario si se actualiza el User
# Si es necesario, usa el siguiente código:
# @receiver(post_save, sender=User)
# def save_usuario(sender, instance, **kwargs):
#     if hasattr(instance, 'usuario'):
#         instance.usuario.save()


"""

# Vistas del Módulo de Inventario



from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Q
from .models import Producto

@permission_required('modulos.view_producto', login_url='inicio')
@login_required
def inventario(request):
    productos = Producto.objects.all()
    query = request.GET.get('q', '')  # Usamos 'q' como parámetro genérico de búsqueda
    
    if query:
        # Filtrar por varios campos usando Q
        productos = productos.filter(
            Q(nombreproducto__icontains=query) |
            Q(nombrecategoria__icontains=query) |
            Q(codigo__icontains=query) |
            Q(calidad__icontains=query) |
            Q(dimension__icontains=query) |
            Q(precio__icontains=query) |
            Q(color__icontains=query) |
            Q(genero__icontains=query)
        )
    
    # Contar el número total de productos para mostrar en la tarjeta
    cantidad_productos = productos.count()

    # Pasamos 'productos' y 'cantidad_productos' al contexto de la plantilla
    return render(request, 'inventario/index.html', {
        'productos': productos,
        'query': query,
        'cantidad_productos': cantidad_productos
    })




@permission_required('modulos.view_insumo', login_url='inicio')
@login_required
def insumo(request):
    query = request.GET.get('q', '')  # Obtener el término de búsqueda de la query
    insumos = Insumo.objects.all()

    if query:
        # Filtrar insumos por múltiples campos usando Q
        insumos = insumos.filter(
            Q(nombreinsumo__icontains=query) |
            Q(color__icontains=query) |
            Q(unidaddemedida__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(dimension__icontains=query)
        )

    cantidad_insumos = insumos.count()  # Contar cuántos insumos existen

    # Pasar la cantidad total de insumos al contexto
    return render(request, 'inventario/insumos.html', {
        'insumos': insumos,
        'query': query,
        'cantidad_insumos': cantidad_insumos
    })



from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render, redirect
from .forms import ProductoForm
from django.contrib.auth.models import User

@permission_required('modulos.add_producto', login_url='inicio')
@login_required
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save(commit=False)  # No guardamos aún el producto
            producto.usuario_idusurios = request.user  # Asignamos el usuario actual al producto
            producto.save()  # Guardamos el producto con el usuario asignado
            return redirect('inventario')
        else:
            print(form.errors)  # Muestra los errores de validación en la consola
    else:
        form = ProductoForm()

    return render(request, 'inventario/crear_producto.html', {'form': form})



@permission_required('modulos.delete_producto', login_url='inicio')
@login_required
def eliminar_producto(request, id):
    registro = get_object_or_404(Producto, idproducto=id)
    registro.delete()
    return redirect('insumo')


@permission_required('modulos.change_producto', login_url='inicio')
@login_required
def editar_producto(request, id):
    producto = get_object_or_404(Producto, idproducto=id)

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)  # Agregar request.FILES
        if form.is_valid():
            form.save()
            return redirect('inventario')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'inventario/editar_producto.html', {'form': form})



@permission_required('modulos.view_compraproducto', login_url='inicio')  # Verificar si el usuario tiene permiso para ver compras de productos
@login_required
def compraProducto(request):
    query = request.GET.get('q', '')  # Capturar el término de búsqueda
    compraproductos = Compraproducto.objects.select_related('producto_idproducto', 'proveedor_idproveedor').all()

    if query:
        # Filtrar por múltiples campos usando Q
        compraproductos = compraproductos.filter(
            Q(fechacompra__icontains=query) |
            Q(producto_idproducto__codigo__icontains=query) |
            Q(producto_idproducto__nombreproducto__icontains=query) |
            Q(proveedor_idproveedor__nombre1__icontains=query) |
            Q(proveedor_idproveedor__apellido1__icontains=query) |
            Q(preciounitario__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(preciototal__icontains=query)
        )

    return render(request, 'inventario/lista_compraProducto.html', {'compraproductos': compraproductos, 'query': query})




@permission_required('modulos.add_compraproducto', login_url='inicio')
@login_required
def crear_compraProducto(request):
    if request.method == 'POST':
        form = CompraProductoForm(request.POST, request.FILES)
        if form.is_valid():
            # Guardamos la compra de producto sin hacer commit
            compra_producto = form.save(commit=False)
            
            # Accedemos al producto relacionado
            producto = compra_producto.producto_idproducto
            
            # Actualizamos la cantidad de productos en el inventario
            producto.cantidad = F('cantidad') + compra_producto.cantidad
            producto.save()

            # Ahora guardamos la compra de producto
            compra_producto.save()
            
            return redirect('inventario')
        else:
            print(form.errors)
    else:
        form = CompraProductoForm()
    
    return render(request, 'inventario/crear_compraProducto.html', {'form': form})



@permission_required('modulos.change_compraproducto', login_url='inicio')
@login_required
def editar_compraProducto(request, id):
    # Obtener la compraProducto original
    compraproducto = get_object_or_404(Compraproducto, idcomprasproductos=id)
    producto = compraproducto.producto_idproducto  # Obtener el producto relacionado
    cantidad_original = compraproducto.cantidad  # Guardar la cantidad original antes de editar

    if request.method == 'POST':
        form = CompraProductoForm(request.POST, request.FILES, instance=compraproducto)  # Agregar request.FILES
        if form.is_valid():
            compra_actualizada = form.save(commit=False)  # Guardar los cambios sin confirmar en la base de datos aún
            nueva_cantidad = compra_actualizada.cantidad  # Nueva cantidad del formulario

            # Calcular la diferencia entre la nueva y la original
            diferencia_cantidad = nueva_cantidad - cantidad_original

            # Actualizar la cantidad del producto en el inventario
            producto.cantidad += diferencia_cantidad
            producto.save()

            # Guardar la compra actualizada
            compra_actualizada.save()

            return redirect('lista_compraProducto')
    else:
        form = CompraProductoForm(instance=compraproducto)  # Mostrar los datos originales del formulario

    return render(request, 'inventario/editar_compraProducto.html', {'form': form})


@permission_required('modulos.delete_compraproducto', login_url='inicio')
@login_required
def eliminar_compraProducto(request, id):
    # Obtener la compraProducto original
    registro = get_object_or_404(Compraproducto, idcomprasproductos=id)
    producto = registro.producto_idproducto  # Obtener el producto relacionado
    cantidad_a_eliminar = registro.cantidad  # Guardar la cantidad que se va a eliminar

    # Actualizar la cantidad del producto en el inventario
    producto.cantidad -= cantidad_a_eliminar  # Restar la cantidad de la compra eliminada
    producto.save()

    # Eliminar la compraProducto
    registro.delete()

    return redirect('lista_compraProducto')





@permission_required('modulos.add_insumo', login_url='inicio')
@login_required
def crear_insumo(request):
    if request.method == 'POST':
        form = InsumoForm(request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            return redirect('insumo')
    else:
        form = InsumoForm()
    
    return render(request, 'inventario/crear_insumo.html', {'form': form})


@permission_required('modulos.change_insumo', login_url='inicio')
@login_required
def editar_insumo(request, id):
    insumo = get_object_or_404(Insumo, idinsumos=id)

    if request.method == 'POST':
        form = InsumoForm(request.POST, request.FILES, instance=insumo)  # Agregar request.FILES
        if form.is_valid():
            form.save()
            return redirect('insumo')
    else:
        form = InsumoForm(instance=insumo)

    return render(request, 'inventario/editar_insumo.html', {'form': form})


@permission_required('modulos.change_comprainsumo', login_url='inicio')
@login_required
def editar_compraInsumo(request, id):
    comprainsumo = get_object_or_404(Comprainsumo, idcomprasinsumos=id)
    insumo = comprainsumo.insumo_idinsumos  # Obtenemos el insumo relacionado

    # Guardamos la cantidad anterior antes de editar
    cantidad_anterior = comprainsumo.cantidad

    if request.method == 'POST':
        form = CompraInsumoForm(request.POST, request.FILES, instance=comprainsumo)
        if form.is_valid():
            # Guardar sin commit para poder manipular la cantidad
            compra_editada = form.save(commit=False)

            # Restar la cantidad anterior del insumo
            insumo.cantidad -= cantidad_anterior

            # Sumar la nueva cantidad al insumo
            insumo.cantidad += compra_editada.cantidad
            insumo.save()  # Guardar la actualización en la cantidad del insumo

            # Guardar la compra de insumo actualizada
            compra_editada.save()

            return redirect('lista_compraInsumo')
    else:
        form = CompraInsumoForm(instance=comprainsumo)

    return render(request, 'inventario/editar_compraInsumo.html', {'form': form})



@permission_required('modulos.delete_comprainsumo', login_url='inicio')
@login_required
def eliminar_compraInsumo(request, id):
    # Obtener el registro de la compra de insumo
    registro = get_object_or_404(Comprainsumo, idcomprasinsumos=id)
    
    # Obtener el insumo relacionado y la cantidad de la compra
    insumo = registro.insumo_idinsumos
    cantidad_comprada = registro.cantidad

    # Restar la cantidad comprada del stock del insumo
    insumo.cantidad -= cantidad_comprada
    insumo.save()  # Guardar los cambios en la tabla Insumo

    # Eliminar la compra de insumo
    registro.delete()

    return redirect('lista_compraInsumo')




@permission_required('modulos.delete_insumo', login_url='inicio')
@login_required
def eliminar_insumo(request, id):
    registro = get_object_or_404(Insumo, idinsumos=id)
    registro.delete()
    return redirect('insumo')


@user_passes_test(es_usuario, login_url='signin')
@login_required
def inicio(request):
    return render(request, 'inventario/inicio.html')



@permission_required('modulos.view_comprainsumo', login_url='inicio')
@login_required
def compraInsumo(request):
    query = request.GET.get('q', '')  # Capturar el término de búsqueda
    comprainsumos = Comprainsumo.objects.select_related('insumo_idinsumos', 'proveedor_idproveedor').all()

    if query:
        # Filtrar por varios campos relacionados y propios usando Q
        comprainsumos = comprainsumos.filter(
            Q(fechacompra__icontains=query) |
            Q(insumo_idinsumos__codigo__icontains=query) |
            Q(insumo_idinsumos__nombreinsumo__icontains=query) |
            Q(proveedor_idproveedor__nombre1__icontains=query) |
            Q(proveedor_idproveedor__apellido1__icontains=query) |
            Q(preciounitario__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(preciototal__icontains=query)
        )

    return render(request, 'inventario/lista_compraInsumo.html', {'comprainsumos': comprainsumos, 'query': query})


@permission_required('modulos.add_comprainsumo', login_url='inicio')
@login_required
def crear_compraInsumo(request):
    if request.method == 'POST':
        form = CompraInsumoForm(request.POST, request.FILES)
        if form.is_valid():
            # Guardar la compra de insumo sin realizar aún el commit
            compra_insumo = form.save(commit=False)
            
            # Obtener el insumo relacionado con la compra
            insumo = compra_insumo.insumo_idinsumos
            
            # Actualizar la cantidad del insumo en inventario
            insumo.cantidad += compra_insumo.cantidad
            insumo.save()  # Guardar el cambio en la cantidad del insumo
            
            # Guardar la compra de insumo en la base de datos
            compra_insumo.save()

            return redirect('insumo')
        else:
            print(form.errors)
    else:
        form = CompraInsumoForm()
    
    return render(request, 'inventario/crear_compraInsumo.html', {'form': form})






@permission_required('modulos.view_proveedor', login_url='inicio')
@login_required
def proveedor(request):
    query = request.GET.get('q', '')  # Captura el término de búsqueda
    proveedores = Proveedor.objects.all()

    if query:
        # Filtrar por varios campos utilizando Q para múltiples condiciones
        proveedores = proveedores.filter(
            Q(nombreempresa__icontains=query) |
            Q(nombre1__icontains=query) |
            Q(nombre2__icontains=query) |
            Q(apellido1__icontains=query) |
            Q(apellido2__icontains=query) |
            Q(tipoproveedor__icontains=query) |
            Q(telefono1__icontains=query) |
            Q(telefono2__icontains=query) |
            Q(categoriaproveedor__icontains=query) |
            Q(direccion__icontains=query) |
            Q(correo1__icontains=query) |
            Q(correo2__icontains=query) |
            Q(sitioweb__icontains=query)
        )

    cantidad_proveedores = proveedores.count()

    return render(request, 'proveedores/index.html', {
        'proveedores': proveedores, 
        'query': query, 
        'cantidad_proveedores': cantidad_proveedores
    })


@permission_required('modulos.add_proveedor', login_url='inicio')
@login_required
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('proveedor')
    else:
        form = ProveedorForm()
    
    return render(request, 'proveedores/crear_proveedores.html', {'form': form})



@permission_required('modulos.change_proveedor', login_url='inicio')
@login_required
def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, idproveedor=id)
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect('proveedor')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'proveedores/editar_proveedores.html', {'form': form})


@permission_required('modulos.delete_proveedor', login_url='inicio')
@login_required
def eliminar_proveedor(request, id):
    registro = get_object_or_404(Proveedor, idproveedor=id)
    registro.delete()
    return redirect('proveedor')


# Vistas del Módulo de Clientes




from django.contrib.auth.decorators import permission_required
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

# Usamos permission_required para asegurarnos que solo usuarios con el permiso 'can_view_clientes' accedan
@permission_required('modulos.can_view_clientes', login_url='inicio')  # 'modulos' es tu app, y 'can_view_clientes' es el permiso
@login_required  # Asegura que el usuario esté autenticado
def clientes(request):
    query = request.GET.get('q', '')  # Si no se pasa un valor, será un string vacío
    clientes = Cliente.objects.all()

    # Filtrar los clientes por varios campos si hay una búsqueda
    if query:
        clientes = clientes.filter(
            Q(nombre1__icontains=query) |
            Q(nombre2__icontains=query) |
            Q(apellido1__icontains=query) |
            Q(apellido2__icontains=query) |
            Q(tipodocumento__icontains=query) |
            Q(numerodocumento__icontains=query) |
            Q(correo1__icontains=query) |
            Q(correo2__icontains=query) |
            Q(telefono1__icontains=query) |
            Q(telefono2__icontains=query) |
            Q(direccion__icontains=query) |
            Q(fechanacimiento__icontains=query) |
            Q(genero__icontains=query)
        )

    # Pasar los resultados y la búsqueda actual al template
    cantidad_clientes = clientes.count()

    
    return render(request, 'clientes/index.html', {
        'clientes': clientes,
        'query': query,
        'cantidad_clientes': cantidad_clientes
    })






@permission_required('modulos.add_cliente', login_url='inicio')
@login_required
def crear_clientes(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('clientes')
    else:
        form = ClienteForm()
    
    return render(request, 'clientes/crear_cliente.html', {'form': form})






@permission_required('modulos.change_cliente', login_url='inicio')
@login_required
def editar_cliente(request, id):
    cliente = get_object_or_404(Cliente, idclientes=id)

    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            return redirect('clientes')
    else:
        form = ClienteForm(instance=cliente)

    return render(request, 'clientes/editar_cliente.html', {'form': form})



@permission_required('modulos.delete_cliente', login_url='inicio')
@login_required
def eliminar_clientes(request, id):
    registro = get_object_or_404(Cliente, idclientes=id)
    registro.delete()
    return redirect('clientes')




@permission_required('modulos.send_email', login_url='inicio') 
@login_required
def enviar_correo(request):
    if request.method == 'POST':
        asunto = request.POST.get('asunto')
        mensaje = request.POST.get('mensaje')
        destinatarios = request.POST.get('destinatarios').split(',')

        try:
            send_mail(
                asunto,
                mensaje,
                'Jvcorporativo@outlook.es',
                destinatarios,
                fail_silently=False,
            )
            return HttpResponseRedirect(reverse('exito'))
        except BadHeaderError:
            return HttpResponse('Encabezado inválido encontrado.')
        except Exception as e:
            return render(request, 'clientes/enviar_correo.html', {
                'error': f"Ocurrió un error al enviar el correo: {str(e)}",
                'destinatarios': request.POST.get('destinatarios'),
            })




def exito(request):
    return render(request, 'clientes/exito.html')




@permission_required('modulos.view_ordendeproduccion', login_url='inicio')
@login_required
def produccion(request):
    query = request.GET.get('q', '')  # Capturar el término de búsqueda
    ordendeproduccion = Ordendeproduccion.objects.all()

    if query:
        # Filtrar por varios campos usando Q
        ordendeproduccion = ordendeproduccion.filter(
            Q(idordendeproduccion__icontains=query) |
            Q(estado__icontains=query) |
            Q(fechacreacion__icontains=query) |
            Q(fechafinalizacion__icontains=query) |
            Q(detalleventa_iddetalleventa__venta_idventas__numeroventa__icontains=query) |
            Q(detalleventa_iddetalleventa__producto_idproducto__codigo__icontains=query) |
            Q(detalleventa_iddetalleventa__producto_idproducto__nombrecategoria__icontains=query) |
            Q(detalleventa_iddetalleventa__producto_idproducto__nombreproducto__icontains=query) |
            Q(detalleventa_iddetalleventa__producto_idproducto__color__icontains=query)
        )

    return render(request, 'produccion/index.html', {'ordendeproduccion': ordendeproduccion, 'query': query})


@permission_required('modulos.add_ordendeproduccion', login_url='inicio')
@login_required
def crear_produccion(request):
    if request.method == 'POST':
        form = OrdendeproduccionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('produccion')
    else:
        form = OrdendeproduccionForm()
    
    return render(request, 'produccion/crear_produccion.html', {'form': form})



@permission_required('modulos.change_ordendeproduccion', login_url='inicio')
@login_required
def editar_produccion(request, id):
    produccion = get_object_or_404(Ordendeproduccion, idordendeproduccion=id)

    if request.method == 'POST':
        form = OrdendeproduccionForm(request.POST, instance=produccion)
        if form.is_valid():
            form.save()
            return redirect('produccion') 
    else:
        form = OrdendeproduccionForm(instance=produccion)

    return render(request, 'produccion/editar_produccion.html', {'form': form})




@permission_required('modulos.view_gastoinsumo', login_url='inicio')
@login_required
def gastoInsumo(request):
    query = request.GET.get('q', '')  # Capturar el término de búsqueda
    controlgastoinsumo = Controlgastoinsumo.objects.select_related('insumo_idinsumos', 'ordendeproduccion_idordendeproduccion').all()

    if query:
        # Filtrar por varios campos usando Q
        controlgastoinsumo = controlgastoinsumo.filter(
            Q(ordendeproduccion_idordendeproduccion__detalleventa_iddetalleventa__venta_idventas__numeroventa__icontains=query) |
            Q(fechauso__icontains=query) |
            Q(insumo_idinsumos__codigo__icontains=query) |
            Q(insumo_idinsumos__nombreinsumo__icontains=query) |
            Q(ordendeproduccion_idordendeproduccion__idordendeproduccion__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(unidaddemedida__icontains=query) |
            Q(comentarios__icontains=query)
        )

    return render(request, 'produccion/lista_gastoInsumo.html', {'controlgastoinsumo': controlgastoinsumo, 'query': query})


@permission_required('modulos.add_controlgastoinsumo', login_url='inicio')
@login_required
def crear_gastoInsumo(request):
    if request.method == 'POST':
        form = GastoInsumoForm(request.POST, request.FILES)
        if form.is_valid():
            # Obtener la instancia del formulario pero no guardar aún
            gasto_insumo = form.save(commit=False)
            
            # Obtener el insumo relacionado y la cantidad utilizada
            insumo = gasto_insumo.insumo_idinsumos
            cantidad_utilizada = gasto_insumo.cantidad

            # Verificar si hay suficiente cantidad en stock
            if insumo.cantidad >= cantidad_utilizada:
                # Restar la cantidad utilizada del stock de insumo
                insumo.cantidad -= cantidad_utilizada
                insumo.save()  # Guardar los cambios en la tabla Insumo
                
                # Guardar el registro del control de gasto de insumo
                gasto_insumo.save()

                return redirect('lista_gastoInsumo')  # Redirigir después de guardar
            else:
                # Manejar el error si no hay suficiente stock
                form.add_error('cantidad', 'No hay suficiente cantidad en stock para este insumo.')

    else:
        form = GastoInsumoForm()

    return render(request, 'produccion/crear_gastoInsumo.html', {'form': form})




@permission_required('modulos.change_controlgastoinsumo', login_url='inicio')
@login_required
def editar_gastoInsumo(request, id):
    controlgastoinsumo = get_object_or_404(Controlgastoinsumo, idcontrolgastoinsumo=id)
    insumo = controlgastoinsumo.insumo_idinsumos  # Obtener el insumo relacionado
    cantidad_original = controlgastoinsumo.cantidad  # Guardar la cantidad original

    if request.method == 'POST':
        form = GastoInsumoForm(request.POST, request.FILES, instance=controlgastoinsumo)
        if form.is_valid():
            gasto_insumo_editado = form.save(commit=False)
            nueva_cantidad = gasto_insumo_editado.cantidad

            # Si la cantidad ha cambiado, ajustar la cantidad del insumo
            diferencia = nueva_cantidad - cantidad_original  # Comparar con la cantidad original

            if insumo.cantidad - diferencia >= 0:
                # Actualizar la cantidad en el insumo
                insumo.cantidad -= diferencia
                insumo.save()  # Guardar los cambios en el insumo

                # Guardar el registro de gasto insumo actualizado
                gasto_insumo_editado.save()

                return redirect('lista_gastoInsumo')
            else:
                form.add_error('cantidad', 'No hay suficiente stock de insumo para cubrir la nueva cantidad.')

    else:
        form = GastoInsumoForm(instance=controlgastoinsumo)

    return render(request, 'produccion/editar_gastoInsumo.html', {'form': form})


@permission_required('modulos.delete_controlgastoinsumo', login_url='inicio')
@login_required
def eliminar_gastoInsumo(request, id):
    registro = get_object_or_404(Controlgastoinsumo, idcontrolgastoinsumo=id)
    insumo = registro.insumo_idinsumos  # Obtener el insumo relacionado
    
    # Restablecer la cantidad del insumo sumando la cantidad del gasto que se va a eliminar
    insumo.cantidad += registro.cantidad
    insumo.save()  # Guardar los cambios en el insumo

    # Eliminar el registro del gasto de insumo
    registro.delete()

    return redirect('lista_gastoInsumo')


# Vistas del Módulo de Stock Ventas

@permission_required('modulos.delete_ordendeproduccion', login_url='inicio')
@login_required
def eliminar_produccion(request, id):
    registro = get_object_or_404(Ordendeproduccion, idordendeproduccion=id)
    registro.delete()
    return redirect('produccion')


# Vistas del Módulo de Ventas
@permission_required('modulos.view_ventas', login_url='inicio')
@login_required
def ventas(request):
    ventas = Venta.objects.all()
    return render(request, 'ventas/index.html', {'ventas': ventas})



@permission_required('modulos.view_ventas', login_url='inicio')
@login_required
def resumen_ventas(request):
    query = request.GET.get('q', '')  # Captura el término de búsqueda
    ventas = Venta.objects.prefetch_related('detalleventa_set__producto_idproducto', 'cliente_idclientes').all()
    ventas_recientes = Venta.objects.filter(modalidadventa='virtual').order_by('-fechaventa')[:5]
   

    if query:
        # Filtrar por varios campos utilizando Q para múltiples condiciones
        ventas = ventas.filter(
            Q(numeroventa__icontains=query) |
            Q(cliente_idclientes__nombre1__icontains=query) |
            Q(cliente_idclientes__nombre2__icontains=query) |
            Q(cliente_idclientes__apellido1__icontains=query) |
            Q(cliente_idclientes__apellido2__icontains=query) |
            Q(cliente_idclientes__tipodocumento__icontains=query) |
            Q(cliente_idclientes__numerodocumento__icontains=query) |
            Q(cliente_idclientes__direccion__icontains=query) |
            Q(valortotal__icontains=query) |
            Q(modalidadventa__icontains=query) |
            Q(estado__icontains=query)
        )

    return render(request, 'ventas/lista_ventas.html', {'ventas': ventas, 'ventas_recientes': ventas_recientes, 'query': query})


@permission_required('modulos.change_venta', login_url='inicio')
@login_required
def anular_venta(request, venta_id):
    if request.method == 'POST':
        venta = get_object_or_404(Venta, idventas=venta_id)
        
        # Verificar si la venta ya está anulada
        if venta.estado == 'Anulada':
            messages.warning(request, 'Esta venta ya ha sido anulada previamente.')
            return redirect('resumen_ventas')
        
        # Obtener todos los detalles de venta relacionados con la venta que se va a anular
        detalles_venta = Detalleventa.objects.filter(venta_idventas=venta)

        # Restaurar las cantidades de productos al inventario
        for detalle in detalles_venta:
            producto = detalle.producto_idproducto
            producto.cantidad += detalle.cantidad  # Restaurar la cantidad de productos al inventario
            producto.save()

        # Cambiar el estado de la venta a 'Anulada'
        venta.estado = 'Anulada'
        venta.save()

        messages.success(request, 'La venta ha sido anulada y el inventario ha sido actualizado.')
        return redirect('resumen_ventas')  # Redirige a la vista donde se muestran las ventas
    else:
        # Manejo de GET si es necesario
        return redirect('resumen_ventas')




# Vista adicional del login (si es necesaria)
def log(request):
    return render(request, 'login.html')


# Vistas adicionales
def home(request):
    return render(request, 'home.html')


@user_passes_test(es_usuario, login_url='signin')
@login_required
def dashboard(request):
    return render(request, 'dashboard.html')


@login_required
def clienteregistrado(request):
    return render(request, 'clientes_pagina/clientes.html')




@permission_required('modulos.view_cliente', login_url='inicio')
@login_required
def seleccionar_clientes_correo(request):
    query = request.GET.get('q', '')  # Obtener el valor de búsqueda de la consulta
    
    # Filtrar los clientes basados en la búsqueda, si se ha ingresado algo
    if query:
        clientes = Cliente.objects.filter(
            Q(nombre1__icontains=query) | 
            Q(nombre2__icontains=query) |
            Q(apellido1__icontains=query) |
            Q(apellido2__icontains=query) |
            Q(numerodocumento__icontains=query) |
            Q(correo1__icontains=query) |
            Q(correo2__icontains=query)
        )
    else:
        clientes = Cliente.objects.all()

    if request.method == 'POST':
        clientes_seleccionados = [c for c in request.POST.getlist('clientes') if c]
        if not clientes_seleccionados:
            return render(request, 'clientes/seleccionar_clientes.html', {
                'clientes': clientes,
                'error': 'Debes seleccionar al menos un cliente para enviar correos.',
                'query': query
            })

        correos = Cliente.objects.filter(idclientes__in=clientes_seleccionados).values('correo1', 'correo2')
        correos_seleccionados = []
        for correo in correos:
            if correo['correo1']:
                correos_seleccionados.append(correo['correo1'])
            if correo['correo2']:
                correos_seleccionados.append(correo['correo2'])
        
        correos_seleccionados = ', '.join(correos_seleccionados)
        
        if correos_seleccionados:
            return redirect('formulario_enviar_correo', correos=correos_seleccionados)
        else:
            return render(request, 'clientes/seleccionar_clientes.html', {
                'clientes': clientes,
                'error': 'No se encontraron correos para los clientes seleccionados.',
                'query': query
            })

    return render(request, 'clientes/seleccionar_clientes.html', {'clientes': clientes, 'query': query})


# Vista para mostrar el formulario de envío de correos
@permission_required('modulos.send_email', login_url='inicio')  # Suponiendo que existe un permiso para enviar correos
@login_required
def formulario_enviar_correo(request, correos=''):
    if request.method == 'POST':
        asunto = request.POST.get('asunto')
        mensaje = request.POST.get('mensaje')
        destinatarios = request.POST.get('destinatarios').split(',') if request.POST.get('destinatarios') else correos.split(',')

        if not asunto or not mensaje:
            return render(request, 'clientes/formulario_correo.html', {
                'destinatarios': ', '.join(destinatarios),
                'error': 'El asunto y el mensaje son obligatorios.'
            })

        try:
            send_mail(
                asunto,
                mensaje,
                'Jvcorporativo@outlook.es',  # Asegúrate de que este correo sea correcto
                destinatarios,
                fail_silently=False,
            )
            return HttpResponseRedirect(reverse('exito'))
        except BadHeaderError:
            return HttpResponse('Encabezado inválido encontrado.')
        except Exception as e:
            return render(request, 'clientes/formulario_correo.html', {
                'error': f"Ocurrió un error al enviar el correo: {str(e)}",
                'destinatarios': ', '.join(destinatarios),
            })
    else:
        return render(request, 'clientes/formulario_correo.html', {'destinatarios': correos})







from django.contrib.auth import login
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.models import Group
from .forms import ClienteSignUpForm
from .models import Cliente
from django.contrib.auth import login
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.models import Group
from .forms import ClienteSignUpForm
from .models import Cliente

from django.contrib.auth import login
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.models import Group
from .models import Cliente
from .forms import ClienteSignUpForm

def registro_cliente(request):
    if request.method == 'POST':
        form = ClienteSignUpForm(request.POST)

        if form.is_valid():
            numero_documento = form.cleaned_data.get('numero_documento')
            email = form.cleaned_data.get('email')

            # Intentamos encontrar un cliente con ese número de documento
            cliente = Cliente.objects.filter(numerodocumento=numero_documento).first()

            if cliente:
                # Si el cliente ya tiene un usuario, mostramos un error
                if cliente.user:
                    messages.error(request, 'Este cliente ya tiene una cuenta de usuario.')
                    return redirect('signin')  # Redirigir al inicio de sesión si ya tiene una cuenta

                # Si el cliente existe pero no tiene usuario, actualizamos su información
                user = form.save(commit=False)
                cliente.user = user
                cliente.correo1 = email
                cliente.nombre1 = form.cleaned_data.get('nombre1')
                cliente.nombre2 = form.cleaned_data.get('nombre2')
                cliente.apellido1 = form.cleaned_data.get('apellido1')
                cliente.apellido2 = form.cleaned_data.get('apellido2')
                cliente.direccion = form.cleaned_data.get('direccion')
                cliente.correo2 = form.cleaned_data.get('correo2')
                cliente.telefono1 = form.cleaned_data.get('telefono1')
                cliente.telefono2 = form.cleaned_data.get('telefono2')
                cliente.fechanacimiento = form.cleaned_data.get('fechanacimiento')
                cliente.genero = form.cleaned_data.get('genero')

                # Guardamos el usuario y el cliente actualizado
                user.save()
                cliente.save()

                # Asignar al usuario al grupo "Clientes"
                cliente_group, created = Group.objects.get_or_create(name='Clientes')
                user.groups.add(cliente_group)

                # Iniciar sesión automáticamente después del registro
                login(request, user, backend='django.contrib.auth.backends.ModelBackend')

                messages.success(request, 'Cuenta creada exitosamente y asociada al cliente existente.')
                return redirect('catalogos')

            else:
                # Si no existe un cliente con ese número de documento, creamos un nuevo cliente
                user = form.save(commit=False)  # Creamos el usuario pero no lo guardamos aún
                user.save()  # Guardamos el usuario

                # Crear un nuevo registro de cliente asociado al usuario
                Cliente.objects.create(
                    user=user,
                    numerodocumento=numero_documento,
                    tipodocumento=form.cleaned_data.get('tipo_documento'),
                    nombre1=form.cleaned_data.get('nombre1'),
                    nombre2=form.cleaned_data.get('nombre2'),
                    apellido1=form.cleaned_data.get('apellido1'),
                    apellido2=form.cleaned_data.get('apellido2'),
                    direccion=form.cleaned_data.get('direccion'),
                    correo1=email,  # Sincronizamos el correo1 con el email
                    correo2=form.cleaned_data.get('correo2'),
                    telefono1=form.cleaned_data.get('telefono1'),
                    telefono2=form.cleaned_data.get('telefono2'),
                    fechanacimiento=form.cleaned_data.get('fechanacimiento'),
                    genero=form.cleaned_data.get('genero')
                )

                # Asignar al usuario al grupo "Clientes"
                cliente_group, created = Group.objects.get_or_create(name='Clientes')
                user.groups.add(cliente_group)

                # Iniciar sesión automáticamente después del registro
                login(request, user, backend='django.contrib.auth.backends.ModelBackend')

                messages.success(request, 'Cliente y usuario creados exitosamente.')
                return redirect('catalogos')

        else:
            messages.error(request, 'Formulario inválido. Por favor, revisa los datos ingresados.')

    else:
        form = ClienteSignUpForm()

    return render(request, 'signup.html', {'form': form})





def asociar_cuenta_cliente(request):
    if request.method == 'POST':
        form = ClienteSignUpForm(request.POST)
        if form.is_valid():
            numero_documento = form.cleaned_data.get('numero_documento')

            try:
                # Buscar al cliente por número de documento
                cliente = Cliente.objects.get(numerodocumento=numero_documento)

                # Si el cliente ya tiene un usuario asociado
                if cliente.user:
                    messages.error(request, 'Este cliente ya tiene una cuenta de usuario.')
                    return redirect('signin')  # Redirigir al inicio de sesión si ya tiene usuario

                # Crear el usuario desde el formulario
                user = form.save(commit=False)
                
                # Guardar el usuario manualmente
                user.save()

                # Asignar al usuario al grupo "Clientes"
                cliente_group, created = Group.objects.get_or_create(name='Clientes')
                user.groups.add(cliente_group)

                # Asociar el usuario recién creado con el cliente
                cliente.user = user
                cliente.save()

                # Iniciar sesión automáticamente
                login(request, user, backend='django.contrib.auth.backends.ModelBackend')

                messages.success(request, 'Tu cuenta ha sido creada exitosamente y se ha asociado al cliente.')
                return redirect('catalogos')  # Redirigir a la página de productos

            except Cliente.DoesNotExist:
                # Si no se encuentra el cliente con el número de documento proporcionado
                messages.error(request, 'No se encontró un cliente registrado con ese número de documento.')
                return redirect('registro_cliente')  # Redirigir al registro de cliente si no se encuentra el cliente

        else:
            messages.error(request, 'Formulario inválido. Por favor, revisa los datos ingresados.')

    else:
        form = ClienteSignUpForm()

    return render(request, 'asociar_cuenta.html', {'form': form})





from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import ActualizarClienteForm
from .models import Cliente

@login_required
def actualizar_datos_cliente(request):
    # Obtener el cliente asociado al usuario actual
    cliente = get_object_or_404(Cliente, user=request.user)

    if request.method == 'POST':
        form = ActualizarClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            # Guardar los cambios en Cliente
            form.save()

            # Si el correo1 se ha actualizado, también actualizar el email del usuario
            if cliente.correo1 != request.user.email:
                request.user.email = cliente.correo1  # Sincronizar con el campo email en User
                request.user.save()

            messages.success(request, 'Tus datos han sido actualizados correctamente.')
            return redirect('home')  # Redirigir a la página de perfil o donde desees
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ActualizarClienteForm(instance=cliente)

    return render(request, 'clientes_pagina/actualizar_datos.html', {'form': form})







# Crear una vista para la verificacion de existencia del cliente



def verifica_cliente(request):
    if request.method == 'POST':
        numero_documento = request.POST.get('numero_documento')  # El número de documento que ingresa el usuario

        # Intentamos encontrar al cliente en la base de datos
        try:
            cliente = Cliente.objects.get(numerodocumento=numero_documento)
            # Si el cliente existe, lo redirigimos a la vista de crear venta, pasando su ID
            return redirect('crear_venta', cliente_id=cliente.idclientes)
        except Cliente.DoesNotExist:
            # Si el cliente no existe, puedes redirigirlo a una vista para crear un nuevo cliente
            return redirect('crear_clientes')  # Asumiendo que tienes una vista para crear clientes

    return render(request, 'ventas/verifica_cliente.html')



# Andres : Descargas 


 # Asumiendo que tu modelo es Compraproducto

@login_required
def descargar_compraproductos(request):
    # Crear el objeto HttpResponse con el tipo de contenido correcto
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="compraproductos.csv"'

    # Crear un escritor CSV
    writer = csv.writer(response)
    writer.writerow(['Fecha', 'Producto', 'Proveedor', 'Precio Unidad', 'Cantidad', 'Precio Total'])

    # Obtener los datos de los productos
    for compra in Compraproducto.objects.all():
        writer.writerow([compra.fechacompra, compra.producto_idproducto.nombreproducto, compra.proveedor_idproveedor, compra.preciounitario, compra.cantidad, compra.preciototal])

    return response



# esto es para descargar ventas
import openpyxl
from django.http import HttpResponse
from .models import Cliente

def exportar_clientes_excel(request):
    # Crear un archivo Excel en memoria
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clientes'

    # Agregar los encabezados de las columnas
    columnas = ['Nombre', 'Apellido', 'Tipo Documento', 'Número Documento', 'Correo 1', 'Correo 2', 'Teléfono 1', 'Teléfono 2', 'Dirección', 'Fecha de Nacimiento', 'Género']
    ws.append(columnas)

    # Obtener los datos de los clientes
    clientes = Cliente.objects.all()

    # Añadir los datos de cada cliente en las filas
    for cliente in clientes:
        ws.append([
            f"{cliente.nombre1} {cliente.nombre2}",
            f"{cliente.apellido1} {cliente.apellido2}",
            cliente.tipodocumento,
            cliente.numerodocumento,
            cliente.correo1,
            cliente.correo2,
            cliente.telefono1,
            cliente.telefono2,
            cliente.direccion,
            cliente.fechanacimiento,
            cliente.genero
        ])

    # Preparar la respuesta para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'
    
    # Guardar el archivo en la respuesta
    wb.save(response)

    return response


def exportar_productos(request):
    # Crear un libro de trabajo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Agregar encabezados a la hoja de Excel
    headers = ['Nombre', 'Categoría', 'Código', 'Calidad', 'Dimensión', 'Precio', 'Color', 'Género', 'Observaciones']
    ws.append(headers)

    # Agregar los datos de los productos a la hoja de Excel
    productos = Producto.objects.all()
    for producto in productos:
        ws.append([
            producto.nombreproducto,
            producto.nombrecategoria,
            producto.codigo,
            producto.calidad,
            producto.dimension,
            producto.precio,
            producto.color,
            producto.genero,
            producto.observaciones
        ])

    # Preparar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'

    # Guardar el archivo de Excel en la respuesta
    wb.save(response)

    return response


import openpyxl
from django.http import HttpResponse
from .models import Insumo  # Asegúrate de importar correctamente el modelo


import openpyxl
from django.http import HttpResponse
from .models import Insumo  # Asegúrate de importar correctamente el modelo
from datetime import datetime

def exportar_insumos(request):
    # Crear un libro de trabajo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Insumos"

    # Agregar encabezados a la hoja de Excel
    headers = ['ID', 'Nombre', 'Color', 'Unidad de Medida', 'Cantidad', 'Fecha de Registro', 'Dimensión']
    ws.append(headers)

    # Agregar los datos de los insumos a la hoja de Excel
    insumos = Insumo.objects.all()
    for insumo in insumos:
        # Si insumo.fecharegistro es un datetime con zona horaria, eliminamos la zona horaria
        if insumo.fecharegistro and isinstance(insumo.fecharegistro, datetime):
            fecha_registro = insumo.fecharegistro.replace(tzinfo=None)
        else:
            fecha_registro = insumo.fecharegistro
        
        ws.append([
            insumo.idinsumos,
            insumo.nombreinsumo,
            insumo.color,
            insumo.unidaddemedida,
            insumo.cantidad,
            fecha_registro,  # Aseguramos que la fecha no tenga información de zona horaria
            insumo.dimension
        ])

    # Preparar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=insumos.xlsx'

    # Guardar el archivo de Excel en la respuesta
    wb.save(response)

    return response


import openpyxl
from django.http import HttpResponse
from .models import Comprainsumo  # Asegúrate de usar el modelo correcto
from datetime import datetime

import openpyxl
from django.http import HttpResponse
from .models import Comprainsumo

def exportar_compras_insumo(request):
    # Crear un libro de trabajo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compra Insumos"

    # Agregar encabezados a la hoja de Excel
    headers = ['Fecha', 'Código', 'Insumo', 'Color', 'Proveedor', 'Precio Unidad', 'Cantidad', 'Precio Total']
    ws.append(headers)

    # Agregar los datos de las compras a la hoja de Excel
    compras = Comprainsumo.objects.all()
    for compra in compras:
        # Manejo de fecha para evitar problemas de timezone
        fecha_compra = compra.fechacompra.replace(tzinfo=None) if compra.fechacompra else ''
        
        # Acceder a las relaciones (insumo y proveedor) con seguridad
        insumo = compra.insumo_idinsumos
        proveedor = compra.proveedor_idproveedor
        
        # Llenar la fila con los datos
        ws.append([
            fecha_compra,
            getattr(insumo, 'codigo', ''),  # Código del insumo
            getattr(insumo, 'nombreinsumo', ''),  # Nombre del insumo
            getattr(compra, 'color', ''),  # Color
            f"{getattr(proveedor, 'nombre1', '')} {getattr(proveedor, 'apellido1', '')}",  # Nombre del proveedor
            compra.preciounitario,  # Precio unitario
            compra.cantidad,  # Cantidad
            compra.preciototal  # Precio total
        ])

    # Preparar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=compras_insumos.xlsx'

    # Guardar el archivo de Excel en la respuesta
    wb.save(response)

    return response



import openpyxl
from django.http import HttpResponse
from .models import Compraproducto

def exportar_compras_productos(request):
    # Crear un libro de trabajo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compra Productos"

    # Agregar encabezados a la hoja de Excel
    headers = ['Fecha', 'Código', 'Producto', 'Proveedor', 'Precio Unidad', 'Cantidad', 'Precio Total']
    ws.append(headers)

    # Agregar los datos de las compras a la hoja de Excel
    compras = Compraproducto.objects.all()  # Asegúrate de usar el modelo correcto
    for compra in compras:
        fecha_compra = compra.fechacompra.replace(tzinfo=None) if compra.fechacompra else ''
        proveedor = compra.proveedor_idproveedor
        ws.append([
            fecha_compra,
            compra.producto_idproducto.codigo,
            compra.producto_idproducto.nombreproducto,
            f"{proveedor.nombre1} {proveedor.apellido1}",
            compra.preciounitario,
            compra.cantidad,
            compra.preciototal
        ])

    # Preparar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=compras_productos.xlsx'

    # Guardar el archivo de Excel en la respuesta
    wb.save(response)

    return response


import openpyxl
from django.http import HttpResponse
from .models import Ordendeproduccion

def exportar_produccion(request):
    # Crear un libro de trabajo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Producción"

    # Agregar encabezados a la hoja de Excel
    headers = ['ID', 'Estado', 'Fecha de Creación', 'Fecha de Finalización', 'Detalle de Venta', 'Usuario']
    ws.append(headers)

    # Agregar los datos de la orden de producción a la hoja de Excel
    producciones = Ordendeproduccion.objects.all()
    for produccion in producciones:
        ws.append([
            produccion.idordendeproduccion,
            produccion.estado,
            produccion.fechacreacion.strftime("%Y-%m-%d %H:%M:%S") if produccion.fechacreacion else '',
            produccion.fechafinalizacion.strftime("%Y-%m-%d %H:%M:%S") if produccion.fechafinalizacion else '',
            produccion.detalleventa_iddetalleventa.iddetalleventa if produccion.detalleventa_iddetalleventa else '',
            produccion.usuario_idusurios.primernombre if produccion.usuario_idusurios else ''
        ])

    # Preparar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=produccion.xlsx'

    # Guardar el archivo de Excel en la respuesta
    wb.save(response)

    return response



import openpyxl
from django.http import HttpResponse
from .models import Controlgastoinsumo

def exportar_gasto_insumos(request):
    # Crear un libro de trabajo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gasto Insumos"

    # Agregar encabezados a la hoja de Excel
    headers = ['Número de Venta', 'Fecha', 'Insumo', 'ID Orden de Producción', 'Cantidad', 'Unidad Medida', 'Comentarios']
    ws.append(headers)

    # Agregar los datos del control de gasto de insumos a la hoja de Excel
    gastos = Controlgastoinsumo.objects.all()
    for gasto in gastos:
        venta_numero = gasto.ordendeproduccion_idordendeproduccion.detalleventa_iddetalleventa.venta_idventas.numeroventa if gasto.ordendeproduccion_idordendeproduccion and gasto.ordendeproduccion_idordendeproduccion.detalleventa_iddetalleventa else ''
        ws.append([
            venta_numero,
            gasto.fechauso.strftime("%Y-%m-%d") if gasto.fechauso else '',
            f"{gasto.insumo_idinsumos.codigo} - {gasto.insumo_idinsumos.nombreinsumo}",
            gasto.ordendeproduccion_idordendeproduccion.idordendeproduccion if gasto.ordendeproduccion_idordendeproduccion else '',
            gasto.cantidad,
            gasto.unidaddemedida,
            gasto.comentarios
        ])

    # Preparar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=gasto_insumos.xlsx'

    # Guardar el archivo de Excel en la respuesta
    wb.save(response)

    return response



import openpyxl
from django.http import HttpResponse
from .models import Venta

def exportar_ventas(request):
    # Crear un nuevo archivo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Agregar los encabezados de las columnas
    headers = [
        'Numero Venta', 'Nombre', 'Apellidos', 'Tipo de Documento',
        'Número de Documento', 'Dirección', 'Productos', 'Fecha',
        'Precio del Producto', 'Modalidad', 'Estado'
    ]
    ws.append(headers)

    # Obtener todos los registros de ventas
    ventas = Venta.objects.all()

    # Agregar los datos de las ventas a la hoja de Excel
    for venta in ventas:
        productos = ", ".join([
            f"Producto: {detalle.producto_idproducto.nombreproducto}, "
            f"Precio: {detalle.producto_idproducto.precio}"
            for detalle in venta.detalleventa_set.all()
        ])

        ws.append([
            venta.numeroventa,
            f"{venta.cliente_idclientes.nombre1} {venta.cliente_idclientes.nombre2}",
            f"{venta.cliente_idclientes.apellido1} {venta.cliente_idclientes.apellido2}",
            venta.cliente_idclientes.tipodocumento,
            venta.cliente_idclientes.numerodocumento,
            venta.cliente_idclientes.direccion,
            productos,
            venta.fechaventa.strftime("%Y-%m-%d"),
            venta.valortotal,
            venta.modalidadventa,
            venta.estado
        ])

    # Preparar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'

    # Guardar el archivo de Excel en la respuesta
    wb.save(response)

    return response

import openpyxl
from django.http import HttpResponse
from .models import Proveedor

def exportar_proveedores(request):
    # Crear un nuevo archivo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    # Agregar los encabezados de las columnas
    headers = [
        'Empresa', 'Nombres', 'Apellidos', 'Tipo proveedor', 'Teléfono', 
        'Teléfono 2', 'Categoría', 'Dirección', 'Correo 1', 'Correo 2', 'Sitio web'
    ]
    ws.append(headers)

    # Obtener todos los proveedores de la base de datos
    proveedores = Proveedor.objects.all()
    # Agregar los datos de los proveedores a la hoja de Excel
    for proveedor in proveedores:
        ws.append([
            proveedor.nombreempresa,
            f"{proveedor.nombre1} {proveedor.nombre2 or ''}",
            f"{proveedor.apellido1} {proveedor.apellido2 or ''}",
            proveedor.tipoproveedor,
            proveedor.telefono1,
            proveedor.telefono2 or '',
            proveedor.categoriaproveedor,
            proveedor.direccion,
            proveedor.correo1,
            proveedor.correo2 or '',
            proveedor.sitioweb or ''
        ])

    # Preparar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'

    # Guardar el archivo de Excel en la respuesta
    wb.save(response)

    return response




# Graficos 




def generar_grafica():
    # Obtener los datos de ventas agrupados por fecha
    ventas_por_fecha = (
        Venta.objects
        .values('fechaventa__date')
        .annotate(total_ventas=Sum('valortotal'))
        .order_by('fechaventa__date')
    )

    # Preparar los datos para la gráfica
    fechas = [str(venta['fechaventa__date']) for venta in ventas_por_fecha]
    valores = [venta['total_ventas'] for venta in ventas_por_fecha]

    # Crear el gráfico
    plt.figure(figsize=(10, 5))
    plt.plot(fechas, valores, marker='o')
    plt.title('Ventas Totales por Fecha')
    plt.xlabel('Fecha')
    plt.ylabel('Total de Ventas')
    plt.xticks(rotation=45)
    plt.tight_layout()

    # Convertir la gráfica a una imagen en formato base64
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    plt.close()
    
    return image_base64

def dashboard_view(request):
    # Generar la gráfica
    grafica = generar_grafica()

    # Obtener otros datos para mostrar en la plantilla
    total_ingresos = Venta.objects.aggregate(Sum('valortotal'))['valortotal__sum'] or 0
    total_ventas = Venta.objects.count()

    contexto = {
        'grafica': grafica,
        'total_ingresos': total_ingresos,
        'total_ventas': total_ventas,
        'total_productos': Producto.objects.count()
    }

    return render(request, 'ventas/graficos.html', contexto)



# Logica para manejar el stock 

from django.shortcuts import get_object_or_404, redirect
from .models import Producto, Compraproducto

def registrar_compra_producto(request):
    if request.method == 'POST':
        producto_id = request.POST.get('producto_id')
        cantidad_comprada = int(request.POST.get('cantidad'))

        # Obtener el producto y actualizar su cantidad
        producto = get_object_or_404(Producto, idproducto=producto_id)
        producto.cantidad += cantidad_comprada
        producto.save()

        # Registrar la compra del producto
        compra = Compraproducto(
            producto_idproducto=producto,
            cantidad=cantidad_comprada,
            preciounitario=request.POST.get('preciounitario'),
            preciototal=request.POST.get('preciototal'),
            fechacompra=request.POST.get('fechacompra'),
            proveedor_idproveedor_id=request.POST.get('proveedor_id')
        )
        compra.save()

        return redirect('lista_compras_productos')  # Redirigir a la lista de compras



def ayuda(request):
    return render(request, 'ayuda.html')
                  

def nosotros(request):
    return render(request, 'nosotros.html')




# Vista para el cliente pueda ver sus compras

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Venta, Detalleventa, Cliente

@login_required
def listar_compras_cliente(request):
    # Obtener el cliente asociado al usuario actual
    cliente = get_object_or_404(Cliente, user=request.user)

    # Obtener todas las compras (ventas) asociadas al cliente
    compras = Venta.objects.filter(cliente_idclientes=cliente).order_by('-fechaventa')

    return render(request, 'clientes_pagina/listar_compras.html', {'compras': compras})



# Todo lo  relacionado con panel de control 



@login_required
def lista_usuarios(request):
    query = request.GET.get('q', '')  # Captura la búsqueda
    usuarios = Usuario.objects.all()

    # Filtrar usuarios por varios campos si hay búsqueda
    if query:
        usuarios = usuarios.filter(
            Q(numerodocumento__icontains=query) |
            Q(primernombre__icontains=query) |
            Q(segundonombre__icontains=query) |
            Q(primeraprellido__icontains=query) |
            Q(segundoaprellido__icontains=query) |
            Q(correo1__icontains=query) |
            Q(correo2__icontains=query) |
            Q(telefono1__icontains=query)
        )

    total_usuarios = usuarios.count()
    # Pasar los resultados filtrados y la búsqueda actual al template
    return render(request, 'panel/lista_usuarios.html', {
        'usuarios': usuarios,
        'query': query,
        'total_usuarios': total_usuarios
    })




from .forms import UsuarioSignUpForm

@login_required
def crear_usuario(request):
    if request.method == 'POST':
        form = UsuarioSignUpForm(request.POST)
        if form.is_valid():
            # Guardar el usuario y asignar rol y permisos
            form.save()
            return redirect('lista_usuarios')
    else:
        form = UsuarioSignUpForm()

    return render(request, 'panel/crear_usuario.html', {'form': form, 'title': 'Crear Usuario'})




@login_required
def editar_usuario(request, pk):
    # Obtener el usuario por su clave primaria (pk)
    usuario = get_object_or_404(Usuario, pk=pk)

    if request.method == 'POST':
        # Llenar el formulario con los datos enviados y el objeto existente
        form = UsuarioSignUpForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect('lista_usuarios')  # Redirigir a la lista de usuarios después de editar
    else:
        # Mostrar el formulario con los datos actuales del usuario
        form = UsuarioSignUpForm(instance=usuario)

    return render(request, 'panel/editar_usuarios.html', {'form': form, 'title': 'Editar Usuario'})



@login_required
def eliminar_usuario(request, pk):
    # Obtener el usuario a eliminar
    usuario = get_object_or_404(Usuario, pk=pk)
    
    # Eliminar el usuario
    usuario.delete()
    
    # Redirigir a la lista de usuarios después de la eliminación
    return redirect('lista_usuarios')




def detalle_venta(request, venta_id):
    venta = get_object_or_404(Venta, idventas=venta_id)
    return render(request, 'ventas/detalle_venta.html', {'venta': venta})




import datetime

def get_motivational_quote():
    # Lista de frases motivadoras
    frases = [
        "¡Hoy es un gran día para aprender algo nuevo!",
        "¡Tú eres capaz de lograr grandes cosas!",
        "Nunca es tarde para comenzar de nuevo.",
        "El éxito no es el destino, es el camino.",
        "¡Sigue adelante, lo mejor está por venir!",
        "Confía en ti mismo, todo es posible.",
        "La perseverancia siempre da frutos.",
        "¡Cada día es una nueva oportunidad!",
        "No te rindas, los grandes logros requieren tiempo.",
        "¡Haz que hoy cuente!"
    ]
    
    # Usamos el día del año para cambiar de frase cada día
    dia_del_anio = datetime.datetime.now().timetuple().tm_yday
    
    # Seleccionamos la frase en función del día del año
    return frases[dia_del_anio % len(frases)]

# Vista
def inicio(request):
    frase_del_dia = get_motivational_quote()
    return render(request, 'inventario/inicio.html', {'frase_del_dia': frase_del_dia})
